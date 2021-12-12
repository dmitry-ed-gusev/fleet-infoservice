#!/usr/bin/env python3
# coding=utf-8

"""
    Scraper for River Register Book.

    Main data source (source system address): https://www.rivreg.ru/

    Useful materials and resources:
      - (search + excel) https://www.rivreg.ru/activities/class/regbook/ - register book url (actual data)
      - (excel direct link) https://www.rivreg.ru/assets/Uploads/Registrovaya-kniga3.xlsx

    Created:  Gusev Dmitrii, 04.05.2021
    Modified: Dmitrii Gusev, 12.12.2021
"""

import logging
from openpyxl import load_workbook
from typing import List

from wfleet.scraper.config import scraper_defaults as const
from wfleet.scraper.utils.utilities import generate_timed_filename
from wfleet.scraper.utils.utilities_http import perform_file_download_over_http
from wfleet.scraper.utils.utilities_xls import process_scraper_dry_run, save_ships_2_excel
from wfleet.scraper.engine.scraper_abstract import ScraperAbstractClass, SCRAPE_RESULT_OK
from wfleet.scraper.entities.ships import ShipDto

# todo: implement unit tests for this module!

RIVER_REG_BOOK_URL = "https://www.rivreg.ru/assets/Uploads/Registrovaya-kniga3.xlsx"

# module logging setup
# log = logging.getLogger(const.SYSTEM_RIVREGRU)
log = logging.getLogger(__name__)
log.debug(f"Logging for module {__name__} is configured.")


def parse_raw_data(raw_data_file: str) -> List[ShipDto]:
    """Parse raw data excel file and return list of ShipDto objects.
    :param raw_data_file:
    :return:
    """
    log.debug(f"Parsing RAW River Register data: {raw_data_file}")

    if raw_data_file is None or len(raw_data_file.strip()) == 0:
        raise ValueError("Provided empty path to raw data!")

    result: List[ShipDto] = list()

    wb = load_workbook(filename=raw_data_file)
    sheet = wb.active

    counter = 0
    for i in range(2, sheet.max_row):  # index rows/cols starts with 1, skip the first row (header)

        # get base key (identity) data for the ship
        imo_number: str = ""
        proprietary_number1: str = sheet.cell(row=i, column=1).value
        proprietary_number2: str = ""

        # skip empty row (won't create empty ship)
        if imo_number is None and proprietary_number1 is None and proprietary_number2 is None:
            log.debug(f"Skipping empty row...")
            continue

        # create new ship object
        ship: ShipDto = ShipDto(imo_number, proprietary_number1, proprietary_number2, const.SYSTEM_RIVREGRU)

        # additional ship info
        ship.main_name = sheet.cell(row=i, column=2).value
        ship.project = sheet.cell(row=i, column=4).value
        ship.build_number = sheet.cell(row=i, column=3).value
        ship.ship_type = sheet.cell(row=i, column=5).value
        ship.build_date = sheet.cell(row=i, column=6).value
        ship.build_place = sheet.cell(row=i, column=7).value

        counter += 1  # increase counter
        log.debug(f"Ship #{counter}: {ship}")

        # add ship to the list
        result.append(ship)

    return result


class RivRegRuScraper(ScraperAbstractClass):
    """Scraper for rivreg.ru source system."""

    def __init__(self, source_name: str, cache_path: str):
        super().__init__(source_name, cache_path)
        self.log = logging.getLogger(const.SYSTEM_RIVREGRU)
        self.log.info(f"RivRegRuScraper: source name {self.source_name}, cache path: {self.cache_path}.")

    def scrap(self, dry_run: bool = False, requests_limit: int = 0):
        """River Register data scraper."""
        log.info("scrap(): processing rivreg.ru")

        if dry_run:  # dry run mode - won't do anything!
            process_scraper_dry_run(const.SYSTEM_RIVREGRU)
            return SCRAPE_RESULT_OK

        # generate scraper cache directory path
        scraper_cache_dir: str = self.cache_path + "/" + generate_timed_filename(self.source_name)

        # download raw data file
        downloaded_file: str = perform_file_download_over_http(RIVER_REG_BOOK_URL, scraper_cache_dir)
        self.log.info(f"Downloaded raw data file: {downloaded_file}")

        # parse raw data into list of ShipDto objects
        ships: List[ShipDto] = parse_raw_data(downloaded_file)
        self.log.info(f"Parsed row data and found {len(ships)} ship(s).")

        excel_file: str = scraper_cache_dir + "/" + const.EXCEL_SHIPS_DATA
        save_ships_2_excel(ships, excel_file)
        self.log.info(f"Found ships saved into {excel_file} file.")

        return SCRAPE_RESULT_OK


# main part of the script
if __name__ == "__main__":
    # print('Don\'t run this script directly! Use wrapper script!')
    parse_raw_data("../cache/02-Jun-2021_17-02-54-scraper_riveregru/Registrovaya-kniga3.xlsx")
