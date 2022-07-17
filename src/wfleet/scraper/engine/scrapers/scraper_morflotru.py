#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
    Scraper for Morflot Ships Book.

    Main data source (source system address): https://morflot.gov.ru/

    Useful materials and resources:
      - *(excel) https://morflot.gov.ru/deyatelnost/transportnaya_bezopasnost/reestryi/reestr_obyektov_i_transportnyih_sredstv/f3926.html
      - (direct link to excel - 02.06.2021) https://morflot.gov.ru/files/docslist/3926-6154-ts_razdel_3.xlsx
      - (direct link to excel - 11.06.2021) https://morflot.gov.ru/files/docslist/3926-5792-ts_razdel_3+.xlsx

    Created:  Dmitrii Gusev, 29.05.2021
    Modified: Dmitrii Gusev, 28.03.2021
"""

import logging
import warnings
from openpyxl import load_workbook
from typing import List
from datetime import datetime
from wfleet.scraper.utils.utilities_http import perform_file_download_over_http
from wfleet.scraper.utils.utilities_xls import process_scraper_dry_run, save_ships_2_excel
from wfleet.scraper.engine.scraper_abstract import ScraperAbstractClass, SCRAPE_RESULT_OK
from wfleet.scraper.config.scraper_config import MSG_MODULE_ISNT_RUNNABLE
from wfleet.scraper.entities.ship import ShipDto

# todo: implement unit tests for this module!
# todo: implement search for new excel file on the page above (see above marker -> *)
# todo: implement processing 'not found' situation (HTTP 404)

# direct URL to excel file
# MORFLOT_DATA_URL = 'http://morflot.gov.ru/files/docslist/3926-6154-ts_razdel_3.xlsx'
# MORFLOT_DATA_URL = 'http://morflot.gov.ru/files/docslist/3926-5792-ts_razdel_3+.xlsx'
MORFLOT_DATA_URL = "http://morflot.gov.ru/files/docslist/3926-4267-ts_razdel_3+.xlsx"

# module logging setup
# log = logging.getLogger(const.SYSTEM_MORFLOTRU)
log = logging.getLogger(__name__)
log.debug(f"Logging for module {__name__} is configured.")


def parse_raw_data(raw_data_file: str) -> List[ShipDto]:
    """Parse raw data excel file and return list of ShipDto objects.
    :param raw_data_file:
    :return:
    """
    log.debug(f"Parsing RAW Morflot data: {raw_data_file}")

    if raw_data_file is None or len(raw_data_file.strip()) == 0:
        raise ValueError("Provided empty path to raw data!")

    result: List[ShipDto] = list()

    wb = load_workbook(filename=raw_data_file)
    sheet = wb.active

    counter = 0
    for i in range(3, sheet.max_row):  # index rows/cols starts with 1, skip the first 2 rows (header)

        # get base key (identity) data for the ship
        imo_number: str = sheet.cell(row=i, column=6).value
        proprietary_number1: str = sheet.cell(row=i, column=7).value
        proprietary_number2: str = sheet.cell(row=i, column=8).value

        # skip empty row (won't create empty ship)
        if imo_number is None and proprietary_number1 is None and proprietary_number2 is None:
            # todo: implement counter for empty rows and report it - too much output!
            log.debug(f"Skipping empty row...")
            continue

        # create new ship object
        ship: ShipDto = ShipDto(imo_number, proprietary_number1, proprietary_number2, SYSTEM_MORFLOTRU)

        # additional ship info
        ship.main_name = sheet.cell(row=i, column=4).value
        ship.home_port = sheet.cell(row=i, column=9).value
        ship.project = sheet.cell(row=i, column=5).value
        ship.owner = sheet.cell(row=i, column=10).value
        ship.owner_address = sheet.cell(row=i, column=11).value
        ship.owner_ogrn = sheet.cell(row=i, column=12).value
        ship.owner_ogrn_date = sheet.cell(row=i, column=13).value

        counter += 1  # increase counter
        log.debug(f"Ship #{counter}: {ship}")

        # add ship to the list
        result.append(ship)

    return result


class MorflotRuScraper(ScraperAbstractClass):
    """Scraper for morflot.ru source system."""

    def __init__(self):
        log.info("MorflotRuScraper: initializing.")

    def scrap(self, dtimestamp: datetime, dry_run: bool, requests_limit: int = 0):
        """Morflot data scraper."""
        log.info("scrap(): processing morflot.ru")

        # not implemented warning
        warnings.warn("This module is not implemented yet!", Warning, stacklevel=2)

        if dry_run:  # dry run mode - won't do anything!
            return SCRAPE_RESULT_OK

        # # generate scraper cache directory path
        # scraper_cache_dir: str = self.cache_path + "/" + generate_timed_filename(self.source_name)
        # # download raw data file
        # downloaded_file: str = perform_file_download_over_http(MORFLOT_DATA_URL, scraper_cache_dir)
        # self.log.info(f"Downloaded raw data file: {downloaded_file}")
        # # parse raw data into list of ShipDto objects
        # ships: List[ShipDto] = parse_raw_data(downloaded_file)
        # self.log.info(f"Parsed row data and found {len(ships)} ship(s).")
        # excel_file: str = scraper_cache_dir + "/" + const.EXCEL_SHIPS_DATA
        # save_ships_2_excel(ships, excel_file)
        # self.log.info(f"Found ships saved into {excel_file} file.")

        return SCRAPE_RESULT_OK


# main part of the script
if __name__ == "__main__":
    print(MSG_MODULE_ISNT_RUNNABLE)
