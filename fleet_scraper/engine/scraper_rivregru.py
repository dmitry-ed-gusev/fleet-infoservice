#!/usr/bin/env python3
# coding=utf-8

"""
    Scraper for River Register Book.

    Main data source (source system address): https://www.rivreg.ru/

    Useful materials and resources:
      - (search + excel) https://www.rivreg.ru/activities/class/regbook/ - register book url (actual data)
      - (excel direct link) https://www.rivreg.ru/assets/Uploads/Registrovaya-kniga3.xlsx

    Created:  Gusev Dmitrii, 04.05.2021
    Modified: Dmitrii Gusev, 11.06.2021
"""

import logging
import shutil
from pathlib import Path
from urllib import request
from openpyxl import load_workbook, Workbook

from fleet_scraper.engine.utils import constants as const
from fleet_scraper.engine.utils.utilities import generate_timed_filename
from fleet_scraper.engine.utils.utilities_http import perform_file_download_over_http
from fleet_scraper.engine.utils.utilities_xls import process_scraper_dry_run
from fleet_scraper.engine.scraper_abstract import ScraperAbstractClass, SCRAPE_RESULT_OK
from fleet_scraper.engine.entities.ships import ShipDto

# todo: implement unit tests for this module!

RIVER_REG_BOOK_URL = 'https://www.rivreg.ru/assets/Uploads/Registrovaya-kniga3.xlsx'

# module logging setup
log = logging.getLogger(const.SYSTEM_RIVREGRU)


def parse_raw_data(raw_data_file: str) -> dict:
    """
    :param raw_data_file:
    :return:
    """
    log.debug(f'Parsing RAW River Register data: {raw_data_file}')

    if raw_data_file is None or len(raw_data_file.strip()) == 0:
        raise ValueError('Provided empty path to raw data!')

    wb = load_workbook(filename=raw_data_file)
    sheet = wb.active

    for i in range(2, sheet.max_row):  # index rows/cols starts with 1, skip the first row (header)
        reg_number = sheet.cell(row=i, column=1).value
        ship: ShipDto = ShipDto('', reg_number, 'rivreg')

        ship.flag = ''
        ship.main_name = sheet.cell(row=i, column=2).value
        ship.secondary_name = ''
        ship.home_port = ''
        ship.call_sign = ''
        ship.extended_info_url = '-'

    return dict()


class RivRegRuScraper(ScraperAbstractClass):
    """Scraper for rivreg.ru source system."""

    def __init__(self, source_name: str, cache_path: str):
        super().__init__(source_name, cache_path)
        self.log = logging.getLogger(const.SYSTEM_RIVREGRU)
        self.log.info(f'RivRegRuScraper: source name {self.source_name}, cache path: {self.cache_path}.')

    def scrap(self, dry_run: bool = False, requests_limit: int = 0):
        """River Register data scraper."""
        log.info("scrap(): processing rivreg.ru")

        if dry_run:  # dry run mode - won't do anything!
            process_scraper_dry_run(const.SYSTEM_RIVREGRU)
            return SCRAPE_RESULT_OK

        # generate scraper cache directory path
        scraper_cache_dir: str = self.cache_path + '/' + generate_timed_filename(self.source_name)

        # download raw data file
        downloaded_file: str = perform_file_download_over_http(RIVER_REG_BOOK_URL, scraper_cache_dir)
        self.log.info(f'Downloaded raw data file: {downloaded_file}')

        return SCRAPE_RESULT_OK


# main part of the script
if __name__ == '__main__':
    # print('Don\'t run this script directly! Use wrapper script!')
    parse_raw_data('./engine/cache/02-Jun-2021_17-02-54-scraper_rivregru/Registrovaya-kniga3.xlsx')
