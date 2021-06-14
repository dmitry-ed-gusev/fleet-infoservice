#!/usr/bin/env python3
# coding=utf-8

"""
    Scraper for Morflot Ships Book.

    Main data source (source system address): http://morflot.gov.ru/

    Useful materials and resources:
      - *(excel) http://morflot.gov.ru/deyatelnost/transportnaya_bezopasnost/reestryi/reestr_obyektov_i_transportnyih_sredstv/f3926.html
      - (direct link to excel - 02.06.2021) http://morflot.gov.ru/files/docslist/3926-6154-ts_razdel_3.xlsx
      - (direct link to excel - 11.06.2021) http://morflot.gov.ru/files/docslist/3926-5792-ts_razdel_3+.xlsx

    Created:  Dmitrii Gusev, 29.05.2021
    Modified: Dmitrii Gusev, 11.06.2021
"""

import logging
import shutil
from pathlib import Path
from urllib import request
from openpyxl import load_workbook, Workbook

from .utils import constants as const
from .utils.utilities import generate_timed_filename
from .utils.utilities_http import perform_file_download_over_http
from .utils.utilities_xls import process_scraper_dry_run
from .scraper_abstract import ScraperAbstractClass, SCRAPE_RESULT_OK
from .entities.ships import BaseShipDto

# todo: implement unit tests for this module!
# todo: implement search for new excel file on the page above (marker as *)

# direct URL to excel file
# MORFLOT_DATA_URL = 'http://morflot.gov.ru/files/docslist/3926-6154-ts_razdel_3.xlsx'
MORFLOT_DATA_URL = 'http://morflot.gov.ru/files/docslist/3926-5792-ts_razdel_3+.xlsx'

# module logging setup
log = logging.getLogger(const.SYSTEM_MORFLOTRU)


def parse_raw_data(raw_data_file: str) -> dict:
    """
    :param raw_data_file:
    :return:
    """
    log.debug(f'Parsing RAW Morflot data: {raw_data_file}')

    if raw_data_file is None or len(raw_data_file.strip()) == 0:
        raise ValueError('Provided empty path to raw data!')

    wb = load_workbook(filename=raw_data_file)
    sheet = wb.active

    for i in range(2, sheet.max_row):  # index rows/cols starts with 1, skip the first row (header)
        reg_number = sheet.cell(row=i, column=1).value
        ship: BaseShipDto = BaseShipDto('', reg_number, 'rivreg')

        ship.flag = ''
        ship.main_name = sheet.cell(row=i, column=2).value
        ship.secondary_name = ''
        ship.home_port = ''
        ship.call_sign = ''
        ship.extended_info_url = '-'

    return dict()


class MorflotRuScraper(ScraperAbstractClass):
    """Scraper for morflot.ru source system."""

    def __init__(self, source_name: str, cache_path: str):
        super().__init__(source_name, cache_path)
        self.log = logging.getLogger(const.SYSTEM_MORFLOTRU)
        self.log.info(f'MorflotRuScraper: source name {self.source_name}, cache path: {self.cache_path}.')

    def scrap(self, dry_run: bool = False, requests_limit: int = 0):
        """Morflot data scraper."""
        log.info("scrap(): processing morflot.ru")

        if dry_run:  # dry run mode - won't do anything!
            process_scraper_dry_run(const.SYSTEM_MORFLOTRU)
            return SCRAPE_RESULT_OK

        # generate scraper cache directory path
        scraper_cache_dir: str = self.cache_path + '/' + generate_timed_filename(self.source_name)

        # download raw data file
        downloaded_file: str = perform_file_download_over_http(MORFLOT_DATA_URL, scraper_cache_dir)
        self.log.info(f'Downloaded raw data file: {downloaded_file}')

        return SCRAPE_RESULT_OK


# main part of the script
if __name__ == '__main__':
    print('Don\'t run this script directly! Use wrapper script!')
